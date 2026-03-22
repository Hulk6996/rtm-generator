"""
excel_report.py — генерация Excel-отчёта RTM.

Листы:
  1. 📊 Дашборд       — KPI, health score, сводка
  2. 📋 RTM            — полная матрица BR → FR → TC
  3. 🔍 Покрытие BR   — покрытие по каждому BR
  4. 👤 По исполнителю — (заглушка, расширяемо)
  5. ⚠ Проблемы       — дубли, конфликты, непокрытые
"""

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import os

from config import COLORS as C, REPORT_TITLE, REPORT_PROJECT, REPORT_VERSION


# ─────────────────────────────────────────────────────────────
#  Style helpers
# ─────────────────────────────────────────────────────────────

def fill(hex_):  return PatternFill("solid", fgColor=hex_)
def fc(argb):    return argb[-6:] if len(argb) == 8 else argb
def font(bold=False, color="000000", size=10, name="Arial"):
    c = fc(color) if len(color) == 8 else color
    return Font(name=name, bold=bold, color=c, size=size)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border(color="FFBDD7EE"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value=None, **kwargs):
    c = ws.cell(row=row, column=col, value=value)
    for attr, val in kwargs.items():
        setattr(c, attr, val)
    return c


def _hdr(ws, row, col, value, bg=C["hdr_bg"], fg=C["hdr_fg"], size=9, h="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = fill(bg)
    c.font      = font(True, fg, size)
    c.alignment = align(h, wrap=True)
    c.border    = border(C["brd"])
    return c


def _body(ws, row, col, value, even=False, h="left", num_fmt=None, color_bg=None, color_fg=None, wrap=False):
    bg = color_bg if color_bg else (C["gry_bg"] if even else C["white"])
    fg = color_fg if color_fg else "000000"
    c  = ws.cell(row=row, column=col, value=value)
    c.fill      = fill(bg)
    c.font      = font(size=9, color=fg)
    c.alignment = align(h, wrap=wrap)
    c.border    = border(C["brd"])
    if num_fmt:
        c.number_format = num_fmt
    return c


# ─────────────────────────────────────────────────────────────
#  Color maps
# ─────────────────────────────────────────────────────────────

RESULT_COLORS = {
    "Passed":  (C["grn_bg"], C["grn_fg"]),
    "Failed":  (C["red_bg"], C["red_fg"]),
    "Blocked": (C["ora_bg"], C["ora_fg"]),
    "Not Run": (C["gry_bg"], C["gry_fg"]),
}
PRIORITY_COLORS = {
    "High":   (C["red_bg"],  C["red_fg"]),
    "Medium": (C["yel_bg"],  C["yel_fg"]),
    "Low":    (C["grn_bg"],  C["grn_fg"]),
}
STATUS_COLORS = {
    "Active":      (C["grn_bg"], C["grn_fg"]),
    "Draft":       (C["yel_bg"], C["yel_fg"]),
    "Rejected":    (C["red_bg"], C["red_fg"]),
    "Deprecated":  (C["gry_bg"], C["gry_fg"]),
}
COV_COLORS = {
    "green":  (C["grn_bg"], C["grn_fg"]),
    "yellow": (C["yel_bg"], C["yel_fg"]),
    "red":    (C["red_bg"], C["red_fg"]),
}


# ─────────────────────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────────────────────

def generate_excel(rtm: dict, metrics: dict, output_path: str) -> None:
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_dashboard(wb, metrics)
    _sheet_rtm(wb, rtm)
    _sheet_br_coverage(wb, rtm, metrics)
    _sheet_issues(wb, metrics)

    wb.save(output_path)
    print(f"  ✔ Excel сохранён: {output_path}")


# ─────────────────────────────────────────────────────────────
#  Sheet 1: Dashboard
# ─────────────────────────────────────────────────────────────

def _sheet_dashboard(wb, metrics):
    ws = wb.create_sheet("📊 ��ашборд")
    ws.sheet_properties.tabColor = "FF1F3864"
    ws.sheet_view.showGridLines   = False
    ws.column_dimensions["A"].width = 2

    col_w = [2, 18, 14, 2, 18, 14, 2, 18, 14, 2]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title row
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 38
    ws.row_dimensions[3].height = 8
    ws.merge_cells("B2:J2")
    c = ws["B2"]
    c.value     = f"{REPORT_TITLE}  |  {REPORT_PROJECT}  v{REPORT_VERSION}"
    c.font      = Font(name="Arial", bold=True, size=16, color=C["hdr_bg"])
    c.alignment = align("left", "center")

    # KPI tiles
    br_cov  = metrics["br_coverage"]
    fr_cov  = metrics["fr_coverage"]
    ts      = metrics["test_stats"]
    health  = metrics["health"]

    kpi = [
        (br_cov['total'],         "BR",              "Всего бизнес-требований"),
        (f"{br_cov['pct_fr']}%",  "Покрытие FR",     "% BR → FR"),
        (f"{fr_cov['pct']}%",     "Покрытие TC",     "% FR → TC"),
        (f"{ts['pct_pass']}%",    "Прошло тестов",   f"{ts['passed']}/{ts['total']}"),
        (ts['failed'],            "Упало тестов",    "Failed"),
        (health["label"],         "Health Score",    f"{health['score']}/100"),
    ]

    tile_pos = [(4,2),(4,5),(4,8),(7,2),(7,5),(7,8)]
    for (row, col), (val, lbl1, lbl2) in zip(tile_pos, kpi):
        ws.row_dimensions[row].height   = 32
        ws.row_dimensions[row+1].height = 18
        ws.merge_cells(start_row=row,   start_column=col, end_row=row,   end_column=col+1)
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
        vc = ws.cell(row=row,   column=col, value=val)
        lc = ws.cell(row=row+1, column=col, value=f"{lbl1} — {lbl2}")
        vc.font      = Font(name="Arial", bold=True, size=22, color=C["acc"])
        vc.alignment = align("center", "center")
        lc.font      = Font(name="Arial", size=9, color=C["gry_fg"])
        lc.alignment = align("center", "center")

    ws.row_dimensions[10].height = 8

    # Status by-priority table
    r0 = 11
    ws.merge_cells(start_row=r0, start_column=2, end_row=r0, end_column=6)
    c = ws.cell(row=r0, column=2, value="Покрытие по приоритету")
    c.fill = fill(C["sub_bg"]); c.font = font(True, C["white"], 11)
    c.alignment = align("left"); c.border = border()

    hdrs = ["Приоритет", "Всего BR", "Покрыто FR", "% FR", "Покрыто TC", "% TC"]
    for ci, h in enumerate(hdrs, 2):
        _hdr(ws, r0+1, ci, h)

    for ri, row_data in enumerate(metrics["priority_cov"]):
        r = r0 + 2 + ri
        even = ri % 2 == 0
        vals = [row_data["priority"], row_data["total"], row_data["with_fr"],
                f"{row_data['pct_fr']}%", row_data["with_tc"], f"{row_data['pct_tc']}%"]
        for ci, v in enumerate(vals, 2):
            _body(ws, r, ci, v, even)
        # colour priority cell
        bg, fg = PRIORITY_COLORS.get(row_data["priority"], (C["white"], "000000"))
        ws.cell(row=r, column=2).fill = fill(bg)
        ws.cell(row=r, column=2).font = font(True, fg, 9)

    # Test results mini-table
    r0b = r0 + len(metrics["priority_cov"]) + 4
    ws.merge_cells(start_row=r0b, start_column=2, end_row=r0b, end_column=5)
    c = ws.cell(row=r0b, column=2, value="Результаты тестирования")
    c.fill = fill(C["sub_bg"]); c.font = font(True, C["white"], 11)
    c.alignment = align("left"); c.border = border()

    for ci, h in enumerate(["Статус", "Кол-во", "%", ""], 2):
        _hdr(ws, r0b+1, ci, h)

    ts_rows = [
        ("Passed",  ts["passed"],  ts["pct_pass"]),
        ("Failed",  ts["failed"],  ts["pct_fail"]),
        ("Blocked", ts["blocked"], round(ts["blocked"]/max(ts["total"],1)*100,1)),
        ("Not Run", ts["not_run"], round(ts["not_run"]/max(ts["total"],1)*100,1)),
    ]
    for ri, (status, cnt, pct) in enumerate(ts_rows):
        r = r0b + 2 + ri
        bg, fg = RESULT_COLORS.get(status, (C["white"], "000000"))
        _body(ws, r, 2, status,  color_bg=bg, color_fg=fg)
        _body(ws, r, 3, cnt, ri%2==0)
        _body(ws, r, 4, f"{pct}%", ri%2==0)


# ─────────────────────────────────────────────────────────────
#  Sheet 2: Full RTM
# ─────────────────────────────────────────────────────────────

def _sheet_rtm(wb, rtm):
    ws = wb.create_sheet("📋 RTM")
    ws.sheet_properties.tabColor = "FF2E75B6"
    ws.sheet_view.showGridLines   = False
    ws.freeze_panes = "A2"

    headers = [
        "BR ID", "BR Название", "Приоритет", "Категория",
        "FR ID", "FR Название", "Тип FR", "Компонент",
        "TC ID", "TC Название", "Тип TC", "Результат",
    ]
    widths = [10, 30, 10, 14, 10, 30, 14, 14, 10, 30, 10, 12]

    for ci, (w, h) in enumerate(zip(widths, headers), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        _hdr(ws, 1, ci, h)
    ws.row_dimensions[1].height = 28

    prev_br = None
    for ri, row in enumerate(rtm["rtm_rows"]):
        r    = ri + 2
        even = ri % 2 == 0

        br_val = row["br_id"] if row["br_id"] != prev_br else ""
        prev_br = row["br_id"]

        vals = [
            row["br_id"],   row["br_title"],  row["br_priority"], row["br_category"],
            row["fr_id"],   row["fr_title"],  row["fr_type"],     row["fr_component"],
            row["tc_id"],   row["tc_title"],  row["tc_type"],     row["tc_result"],
        ]
        for ci, v in enumerate(vals, 1):
            _body(ws, r, ci, v, even)

        # Colorize result
        if row["tc_result"]:
            bg, fg = RESULT_COLORS.get(row["tc_result"], (C["white"], "000000"))
            ws.cell(row=r, column=12).fill = fill(bg)
            ws.cell(row=r, column=12).font = font(size=9, color=fg)

        # Colorize priority
        if row["br_priority"]:
            bg, fg = PRIORITY_COLORS.get(row["br_priority"], (C["white"], "000000"))
            ws.cell(row=r, column=3).fill = fill(bg)
            ws.cell(row=r, column=3).font = font(size=9, color=fg)

        # Highlight empty FR (BR without FR)
        if not row["fr_id"]:
            ws.cell(row=r, column=5).fill = fill(C["red_bg"])
            ws.cell(row=r, column=5).value = "⚠ Нет FR"

        # Highlight empty TC
        if row["fr_id"] and not row["tc_id"]:
            ws.cell(row=r, column=9).fill = fill(C["yel_bg"])
            ws.cell(row=r, column=9).value = "⚠ Нет TC"

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ─────────────────────────────────────────────────────────────
#  Sheet 3: BR Coverage
# ─────────────────────────────────────────────────────────────

def _sheet_br_coverage(wb, rtm, metrics):
    ws = wb.create_sheet("🔍 Покрытие BR")
    ws.sheet_properties.tabColor = "FF375623"
    ws.sheet_view.showGridLines   = False
    ws.freeze_panes = "A2"

    br_map    = rtm["br_map"]
    br_to_frs = rtm["br_to_frs"]
    fr_to_tcs = rtm["fr_to_tcs"]

    headers = ["BR ID", "Название", "Приоритет", "Категория",
               "FR (кол-во)", "TC (кол-во)", "% Покрытие FR", "% Покрытие TC", "Статус"]
    widths  = [10, 34, 10, 14, 12, 12, 14, 14, 16]

    for ci, (w, h) in enumerate(zip(widths, headers), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        _hdr(ws, 1, ci, h)
    ws.row_dimensions[1].height = 28

    for ri, (br_id, br) in enumerate(br_map.items()):
        r      = ri + 2
        even   = ri % 2 == 0
        frs    = br_to_frs.get(br_id, [])
        tcs_   = []
        for f in frs:
            tcs_.extend(fr_to_tcs.get(f, []))
        tcs_unique = list(dict.fromkeys(tcs_))

        pct_fr = 100.0 if frs else 0.0
        pct_tc = 100.0 if tcs_unique else 0.0
        status_ = "✅ Полное" if frs and tcs_unique else ("⚠ Частичное" if frs else "❌ Нет покрытия")

        vals = [br_id, br["title"], br["priority"], br["category"],
                len(frs), len(tcs_unique), f"{pct_fr:.0f}%", f"{pct_tc:.0f}%", status_]
        for ci, v in enumerate(vals, 1):
            _body(ws, r, ci, v, even)

        # Colour priority
        bg, fg = PRIORITY_COLORS.get(br["priority"], (C["white"], "000000"))
        ws.cell(row=r, column=3).fill = fill(bg)
        ws.cell(row=r, column=3).font = font(size=9, color=fg)

        # Colour status
        if "Полное" in status_:
            ws.cell(row=r, column=9).fill = fill(C["grn_bg"])
            ws.cell(row=r, column=9).font = font(size=9, color=C["grn_fg"])
        elif "Частичное" in status_:
            ws.cell(row=r, column=9).fill = fill(C["yel_bg"])
            ws.cell(row=r, column=9).font = font(size=9, color=C["yel_fg"])
        else:
            ws.cell(row=r, column=9).fill = fill(C["red_bg"])
            ws.cell(row=r, column=9).font = font(size=9, color=C["red_fg"])

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ─────────────────────────────────────────────────────────────
#  Sheet 4: Issues
# ─────────────────────────────────────────────────────────────

def _sheet_issues(wb, metrics):
    ws = wb.create_sheet("⚠ Проблемы")
    ws.sheet_properties.tabColor = "FFFF0000"
    ws.sheet_view.showGridLines   = False
    ws.freeze_panes = "A2"

    headers = ["Тип", "ID 1", "ID 2 / Описание", "Детали"]
    widths  = [22, 12, 30, 50]
    for ci, (w, h) in enumerate(zip(widths, headers), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        _hdr(ws, 1, ci, h)

    rows = []

    for item in metrics["uncovered_br"]:
        rows.append(("❌ BR без FR",    item, "—", "Бизнес-требование не связано ни с одним FR"))
    for item in metrics["uncovered_fr"]:
        rows.append(("❌ FR без TC",    item, "—", "Функциональное требование не покрыто тест-кейсом"))
    for item in metrics["orphan_fr"]:
        rows.append(("⚠ FR без BR",    item, "—", "FR не привязан ни к одному бизнес-требованию"))
    for item in metrics["orphan_tc"]:
        rows.append(("⚠ TC без FR",    item, "—", "Тест-кейс не привязан ни к одному FR"))
    for dup in metrics["duplicates"]:
    0   rows.append(("🔁 Дубль",       dup["id1"], dup["id2"], f"{dup['reason']}: «{dup['title']}»"))
    for conf in metrics["conflicts"]:
        rows.append(("⚡ Конфликт",    conf["id1"], conf["id2"], f"{conf['reason']} [категория: {conf['category']}]"))

    if not rows:
        ws.cell(row=2, column=1, value="✅ Проблем не обнаружено")
        ws.cell(row=2, column=1).font = font(True, C["grn_fg"], 11)
        return

    for ri, (typ, id1, id2, detail) in enumerate(rows):
        r    = ri + 2
        even = ri % 2 == 0
        _body(ws, r, 1, typ,    even)
        _body(ws, r, 2, id1,    even)
        _body(ws, r, 3, id2,    even)
        _body(ws, r, 4, detail, even, wrap=False)

        # Colour by type
        if "❌" in typ:
            bg, fg = C["red_bg"], C["red_fg"]
        elif "⚠" in typ:
            bg, fg = C["yel_bg"], C["yel_fg"]
        elif "🔁" in typ:
            bg, fg = C["ora_bg"], C["ora_fg"]
        else:
            bg, fg = C["sub_bg"], C["sub_fg"]
        ws.cell(row=r, column=1).fill = fill(bg)
        ws.cell(row=r, column=1).font = font(True, fg, 9)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
